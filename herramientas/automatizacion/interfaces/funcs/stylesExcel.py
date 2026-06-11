import tkinter.messagebox as mb
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter as getColLetter
from openpyxl.cell.cell import MergedCell

def applyStyleExcel(pathFile):
    """
    Aplica estilos visuales (colores de cabecera, bordes, fuentes, auto-filtros y anchos) a un Excel.
    
    Aplica formateo diferenciado según el nombre de la pestaña (ej. 'Reporte', 'Reporte SLA', 'Eventos', etc.).
    Ajusta el ancho de las columnas y asigna colores semánticos (verde para a tiempo, rojo para vencido).
    
    Args:
        pathFile (str): Ruta al archivo Excel en el disco que se va a formatear.
        
    Raises:
        ValueError: Si el archivo Excel está vacío o ocurre un error durante el procesamiento de los estilos.
    """
    
    try:
        wb = openpyxl.load_workbook(pathFile)
        if not wb.sheetnames:
            raise ValueError("El archivo Excel esta vacio, no se aplican estilos.")
        
        # Paleta de colores y estilos de fuente corporativos
        blueFill = PatternFill(start_color="1B7EF7", fill_type="solid")
        orangeFill = PatternFill(start_color="FAAA15", fill_type="solid")
        redFill = PatternFill(start_color="FF0000", fill_type="solid")
        greenFill = PatternFill(start_color="00FF00", fill_type="solid")
        yellowFill = PatternFill(start_color="FFFF00", fill_type="solid")
        blackFill = PatternFill(start_color="000000", fill_type="solid")
        grayFill = PatternFill(start_color="6E6D6D", fill_type="solid")
        totalFill = PatternFill(start_color="B6C1FF", fill_type='solid')
        whiteFont = Font(color="FFFFFF", bold=True, name="Arial")
        centerAlignment = Alignment(horizontal='center', vertical='center')
        borderStyle = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Iterar por cada pestaña del libro
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            
            # Aplicar estilos a las cabeceras y auto-filtros en hojas de datos planos
            if sheet in ['Reporte', 'Datos', 'insumoCMDB', 'Datos CRQ', 'Datos WO']:
                ws.auto_filter.ref = f'A1:{getColLetter(ws.max_column)}1'
                for cell in ws[1]:
                    cell.fill = blueFill
                    cell.font = whiteFont
                    cell.border = borderStyle
                    cell.alignment = centerAlignment
                    ws.column_dimensions[cell.column_letter].width = 35
            
            # Aplicar colores semánticos a las celdas del reporte SLA
            elif sheet == 'Reporte SLA':
                ws.auto_filter.ref = f'A1:{getColLetter(ws.max_column)}1'
                colCumplimiento = []
                for cell in ws[1]:
                    cell.fill = orangeFill
                    cell.font = Font(color="000000", bold=True, name='Arial')
                    cell.border = borderStyle
                    cell.alignment = centerAlignment
                    if cell.value in ['Cumplimiento SLA 4.1', 'Cumplimiento SLA 4.2', 'Observaciones']:
                        colCumplimiento.append(cell.column_letter)
                    ws.column_dimensions[cell.column_letter].width = 25
                
                # Resaltar en verde las celdas "A Tiempo", en rojo las "Vencido" y en amarillo los registros "Por validar"
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    for col in colCumplimiento:
                        cell = ws[f'{col}{row[0].row}']
                        if cell.value == 'A Tiempo':
                            cell.fill = greenFill
                            cell.font = Font(color='000000')
                        elif cell.value == 'Vencido':
                            cell.fill = redFill
                            cell.font = Font(color='FFFFFF')
                        elif cell.value == 'Por validar':
                            for column in range(1, ws.max_column + 1):
                                ws.cell(row=cell.row, column=column).fill = yellowFill

            # Formatear pestañas complejas con tablas de resumen y celdas combinadas
            elif sheet in ['Eventos', 'Incidentes']:
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                    for cell in row:
                        if cell.row == 1:
                            ws.column_dimensions[cell.column_letter].width = 20
                        if not isinstance(cell, MergedCell):
                            if not (cell.value is None):
                                cell.border= borderStyle
                                # Detectar títulos de resumen y colorear filas totales
                                if cell.value == 'Resumen SLA por Torre':
                                    cell.fill = blackFill
                                    cell.font = whiteFont
                                    for i in range(cell.row, ws.max_row + 1):
                                        cellAdjacent = ws.cell(row=i, column=cell.column)
                                        if cellAdjacent.value == 'Total':
                                            for j in range(1, ws.max_column + 1):
                                                cellAdjacent = ws.cell(row=i, column=j)
                                                if not (cellAdjacent.value in ['', None]):
                                                    cellAdjacent.fill = totalFill
                                                    letter = getColLetter(j)
                                            break
                                    ws.merge_cells(f'A{cell.row}:{letter}{cell.row}')
                                    
                                elif cell.value in ['Grupos con eventos vencidos', 'Grupos con incidentes vencidos']:
                                    cell.fill = blackFill
                                    cell.font = whiteFont
                                    ws.merge_cells(f'A{cell.row}:B{cell.row}')
                                elif cell.value == 'Grupo Asignado':
                                    cell.fill = grayFill
                                    cell.font = whiteFont
                               
        wb.save(pathFile)
        wb.close()
    except ValueError as ve:
        mb.showerror(title="Error", message=ve)
        
    except Exception as e:
        print(f'Error al aplicar estilos al archivo Excel: {e}')
        raise ValueError("Error al aplicar estilos al archivo Excel")
        
if __name__ == "__main__":
    path = 'C:\\Users\\2898604\\Desktop\\reporteSLA.xlsx'
    applyStyleExcel(path)
